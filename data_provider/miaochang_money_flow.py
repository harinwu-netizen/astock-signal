"""
资金流数据获取 - 妙查 fallback
通过 subprocess 调用 mx-finance-data skill,解析 Excel 提取关键字段
"""

import json
import logging
import os
import subprocess
from typing import Any, Dict, Optional

logger = logging.getLogger("MoneyFlow.MiaoChang")

MX_SCRIPT = "/root/.openclaw/workspace/skills/mx-finance-data/scripts/get_data.py"
# 妙查脚本输出目录(以 cwd 为基准)
# 信号灯从 astock-signal/ 跑,所以这里是 astock-signal/miaoxiang/mx_finance_data
MX_OUTPUT_DIR = "miaoxiang/mx_finance_data"


def _get_latest_xlsx_for_code(code: str) -> Optional[str]:
    """
    从妙查输出目录找最新生成的 xlsx 文件，
    优先匹配:description.txt 的"查询内容"行包含本股票代码
    """
    if not os.path.isdir(MX_OUTPUT_DIR):
        return None
    all_files = [
        os.path.join(MX_OUTPUT_DIR, f)
        for f in os.listdir(MX_OUTPUT_DIR)
        if f.endswith(".xlsx")
    ]
    if not all_files:
        return None

    candidates = []
    for xlsx in all_files:
        desc_path = xlsx.replace(".xlsx", "_description.txt")
        if not os.path.exists(desc_path):
            continue
        try:
            with open(desc_path, encoding="utf-8") as f:
                lines = f.readlines()
            # 妙查描述第 3 行格式:"查询内容: 000629钒钛股份今日资金流向..."
            query_line = next((l for l in lines if l.startswith("查询内容:")), "")
            if code in query_line:
                candidates.append((os.path.getmtime(xlsx), xlsx))
        except Exception:
            pass
    if candidates:
        candidates.sort(reverse=True)
        return candidates[0][1]

    # 退化:拿最新文件
    return max(all_files, key=os.path.getmtime)


def _clean_value(v: Any) -> Optional[float]:
    """清洗值,去掉'元''万''亿'等单位后缀,转数字"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    for unit in ("元", "万", "亿", "%", "万元", "亿元"):
        s = s.replace(unit, "")
    s = s.strip().replace("+", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_xlsx(xlsx_path: str) -> Dict[str, Any]:
    """
    解析妙查输出的 xlsx,提取资金流关键字段
    取所有 sheet 中第一个出现的字段值(避免 sheet 重复时覆盖)
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result: Dict[str, Any] = {"_sheets": wb.sheetnames}

    # 关注的字段(妙查输出里可能的中文 key)
    keys_of_interest = {
        "最新价", "收盘价",
        "主力净额", "主力净流入", "主力净流入资金",
        "超大单净额", "超大单净流入资金",
        "大单净额", "大单净流入资金",
        "中单净额", "小单净额",
        "主力流入", "主力流入资金", "主力流出", "主力流出资金",
        "当日DDX", "当日DDY",
        "3日DDX", "5日DDX", "10日DDX",
    }

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if key in keys_of_interest and key not in result:
                result[key] = _clean_value(row[1])

    return result


def get_money_flow_via_miaochang(code: str, name: str = "",
                                 timeout: int = 60) -> Optional[Dict[str, Any]]:
    """
    通过妙查获取单只股票资金流

    Args:
        code: 股票代码,如 "000629"
        name: 股票名称,如 "钒钛股份"
        timeout: 子进程超时(秒)

    Returns:
        dict 或 None:
            {
                "code": str, "name": str,
                "main_net": float(万), "main_in": float(万), "main_out": float(万),
                "big_net": float(万), "super_net": float(万),
                "small_net": float(万), "mid_net": float(万),
                "ddx": float, "ddy": float, "date": str,
            }
    """
    query = f"{code}{name}今日资金流向，主力净流入，最新股价，DDX"
    logger.info(f"[MC] 调用妙查: code={code} name={name}")

    try:
        # 1. 记录调用前最新的 xlsx 文件(用来跳过旧的)
        existing_files = set()
        if os.path.isdir(MX_OUTPUT_DIR):
            existing_files = set(os.listdir(MX_OUTPUT_DIR))

        # 2. 调妙查脚本
        result = subprocess.run(
            ["python3", MX_SCRIPT, "--query", query],
            capture_output=True, text=True, timeout=timeout,
        )
        if result.returncode != 0:
            logger.warning(f"[MC] 妙查脚本失败: {result.stderr[:200]}")
            return None

        # 3. 找调用后新生成的 xlsx(优先匹配本股票代码)
        new_files = set(os.listdir(MX_OUTPUT_DIR)) - existing_files
        xlsx_files = [f for f in new_files if f.endswith(".xlsx")]
        if xlsx_files:
            # 有新文件:在新增文件里找 code 匹配的
            new_full = [os.path.join(MX_OUTPUT_DIR, f) for f in xlsx_files]
            matched = []
            for x in new_full:
                desc = x.replace(".xlsx", "_description.txt")
                if os.path.exists(desc):
                    try:
                        with open(desc, encoding="utf-8") as f:
                            lines = f.readlines()
                        query_line = next((l for l in lines if l.startswith("查询内容:")), "")
                        if code in query_line:
                            matched.append(x)
                    except Exception:
                        pass
            if matched:
                latest = max(matched, key=os.path.getmtime)
            else:
                latest = max(new_full, key=os.path.getmtime)
        else:
            # 无新文件:全局找 code 匹配
            latest = _get_latest_xlsx_for_code(code)

        if not latest or not os.path.exists(latest):
            logger.warning(f"[MC] 未找到妙查输出文件")
            return None

        # 4. 解析
        data = _parse_xlsx(latest)
        if not data or ("主力净额" not in data and "主力净流入" not in data
                        and "主力净流入资金" not in data):
            logger.warning(f"[MC] 妙查返回数据缺少关键字段: {data.keys()}")
            return None

        # 5. 字段归一化
        return {
            "code": code,
            "name": name,
            "main_net": data.get("主力净流入") or data.get("主力净流入资金")
                          or data.get("主力净额") or 0.0,
            "main_in": data.get("主力流入资金") or data.get("主力流入") or 0.0,
            "main_out": data.get("主力流出资金") or data.get("主力流出") or 0.0,
            "big_net": data.get("大单净额") or data.get("大单净流入资金") or 0.0,
            "big_in": 0.0,  # 妙查不区分大单流入/流出
            "big_out": 0.0,
            "super_net": data.get("超大单净额") or data.get("超大单净流入资金") or 0.0,
            "super_in": 0.0,
            "super_out": 0.0,
            "small_net": data.get("小单净额") or 0.0,
            "mid_net": data.get("中单净额") or 0.0,
            "ddx": data.get("当日DDX") or 0.0,
            "ddy": data.get("当日DDY") or 0.0,
            "date": "",  # 妙查结果里有日期,但解析复杂,留给上层按需补
        }

    except subprocess.TimeoutExpired:
        logger.warning(f"[MC] 妙查超时({timeout}s): code={code}")
        return None
    except Exception as e:
        logger.error(f"[MC] 妙查异常: {e}")
        return None
