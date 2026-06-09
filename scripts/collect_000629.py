#!/usr/bin/env python3
"""
000629 攀钢钒钛 · 每日自动采集脚本

v6.11 (2026-06-09): 数据源切换 push2delay + txstock (避免妙查 403 限流)
  - 资金流: push2delay (免费, 几乎不限流) — 推后查讯列
  - 行情:   txstock (腾讯财经) — 免费, 包括开/高/低/昨收
  - DDX:    妙查 独有, push2delay 不提供, 标记 N/A
  - fallback: 妙查 (可继续使用但默认不再走)

功能：
1. 拉今日 000629 行情 + 资金流 (push2delay) + DDX (妙查 fallback)
2. 写入 memory/YYYY-MM-DD.md (沿用现有格式)
3. 写 data/outbox/ 由 outbox_daemon 转发到飞书
4. 写 data/send_health.json 健康检查

用法：
  python3 collect_000629.py              # 拉今日数据
  python3 collect_000629.py --dry-run    # 拉数据但不写文件
  python3 collect_000629.py --source push2delay  # 指定数据源
"""

import sys
import os
import json
import subprocess
import argparse
from datetime import datetime
from pathlib import Path

# 项目根
PROJECT_ROOT = Path("/root/.openclaw/workspace")
MX_FINANCE_SCRIPT = PROJECT_ROOT / "skills/mx-finance-data/scripts/get_data.py"
FUND_FLOW_SCRIPT_DIR = PROJECT_ROOT / "scripts"
TXSTOCK_SCRIPT_DIR = PROJECT_ROOT / "skills/txstock/scripts"

# v6.11: 数据源优先级
SOURCE_PUSH2DELAY = "push2delay"
SOURCE_TXSTOCK = "txstock"
SOURCE_MIAOCHANG = "miaochang"  # fallback


def fetch_push2delay_data(stock_code: str = "000629", stock_name: str = "钒钛股份") -> dict:
    """
    v6.11: 从东方财富 push2delay 公开端点拉资金流
    返回字段: main_net / super_net / big_net / mid_net / small_net (单位: 万元)
             close / change_pct / main_pct 等
    """
    try:
        sys.path.insert(0, str(FUND_FLOW_SCRIPT_DIR))
        from fund_flow_api import fetch_public_flow
        pub = fetch_public_flow(stock_code, stock_name)
        if not pub:
            return {"error": "push2delay 返回 None"}
        # 映射到妙查字段名以兼容 format_daily_log
        return {
            "主力净流入": pub.get("main_net"),
            "超大单净额": pub.get("super_net"),
            "大单净额": pub.get("big_net"),
            "中单净额": pub.get("mid_net"),
            "小单净额": pub.get("small_net"),
            "收盘价": pub.get("close"),
            "涨跌幅": pub.get("change_pct"),
            "主力净流入资金": pub.get("main_net"),  # 别名
            "_source": "push2delay",
        }
    except Exception as e:
        return {"error": f"push2delay 异常: {e}"}


def fetch_txstock_data(stock_code: str = "000629") -> dict:
    """
    v6.11: 从腾讯财经拉行情数据 (开/高/低/昨收)
    txstock 返回价格字段是元
    """
    try:
        sys.path.insert(0, str(TXSTOCK_SCRIPT_DIR))
        from txstock import TxStock
        ts = TxStock()
        rt = ts.get_realtime(stock_code)
        if not rt or "error" in rt:
            return {"error": f"txstock: {rt.get('error', '返回空')}"}
        return {
            "开盘": rt.get("open"),
            "最高": rt.get("high"),
            "最低": rt.get("low"),
            "昨收": rt.get("prev_close"),
            "收盘价": rt.get("price"),  # 收盘以腾讯为准
            "涨跌幅": rt.get("change_pct"),
            "_source": "txstock",
        }
    except Exception as e:
        return {"error": f"txstock 异常: {e}"}


def fetch_000629_data(stock_code: str = "000629", stock_name: str = "钒钛股份") -> dict:
    """
    v6.11: 多源拉取 000629 今日数据
    优先级: push2delay (资金流) + txstock (行情) → 妙查 (全部 fallback)
    返回合并后的 dict, 字段名与妙查原输出一致
    """
    result = {}

    # Step 1: push2delay 拉资金流 (免费, 几乎不限流)
    pub = fetch_push2delay_data(stock_code, stock_name)
    if "error" in pub:
        print(f"⚠️  push2delay 失败: {pub['error']}, 尝试 txstock + 妙查")
    else:
        result.update(pub)
        print(f"✅ push2delay 资金流: 主力={pub.get('主力净流入')}万")

    # Step 2: txstock 拉行情 (免费, 包括开/高/低/昨收)
    ts_data = fetch_txstock_data(stock_code)
    if "error" in ts_data:
        print(f"⚠️  txstock 失败: {ts_data['error']}")
        # push2delay 有 close/change_pct, 使用之
        if "收盘价" not in result:
            return {"error": f"所有数据源都失败: push2delay={pub.get('error')}, txstock={ts_data.get('error')}"}
    else:
        # 合并行情, 优先使用 txstock (包括昨收)
        for k in ("开盘", "最高", "最低", "昨收", "收盘价", "涨跌幅"):
            if ts_data.get(k) is not None:
                result[k] = ts_data[k]
        print(f"✅ txstock 行情: 收盘={ts_data.get('收盘价')} 开={ts_data.get('开盘')}")

    # Step 3: DDX/DDY 仍需妙查 (push2delay 不提供)
    # 默认 N/A
    result.setdefault("当日DDX", "N/A")
    result.setdefault("3日DDX", "N/A")
    result.setdefault("5日DDX", "N/A")
    result.setdefault("10日DDX", "N/A")
    result["_ddx_source"] = "unavailable (push2delay 不提供 DDX)"

    if not result:
        return {"error": "所有数据源均失败"}

    return result


def fetch_miaochang_data(stock_code: str = "000629", stock_name: str = "钒钛股份") -> dict:
    """
    调用妙查脚本拉数据,解析 Excel 提取关键字段 (v6.11 后只作 fallback)
    """
    query = f"{stock_code}{stock_name}今日资金流向，主力净流入，最新股价"
    cmd = ["python3", str(MX_FINANCE_SCRIPT), "--query", query]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        if result.returncode != 0:
            return {"error": f"妙查脚本失败: {result.stderr[:300]}"}

        # 解析输出:获取文件路径
        output = result.stdout
        # 找 xlsx 文件路径
        xlsx_path = None
        for line in output.split("\n"):
            if ".xlsx" in line and ":" in line:
                # 行格式:文件: /path/to/file.xlsx
                xlsx_path = line.split(":")[-1].strip()
                break

        if not xlsx_path or not os.path.exists(xlsx_path):
            return {"error": f"未找到 Excel 输出"}

        # 读取 Excel
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        # 提取关键字段
        data = {"raw_sheets": wb.sheetnames}

        def _clean_value(v):
            """清洗值,去掉'元''万''亿'等单位后缀,转数字"""
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return v
            s = str(v).strip()
            # 去掉单位后缀
            for unit in ("元", "万", "亿", "%", "万元", "亿元"):
                s = s.replace(unit, "")
            s = s.strip()
            # 尝试转数字
            try:
                return float(s.replace("+", ""))
            except (ValueError, TypeError):
                return s if s else None

        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                key = str(row[0]).strip()
                if key in ("最新价", "收盘价", "开盘", "最高", "最低", "昨收",
                           "主力净额", "主力净流入", "主力净流入资金",
                           "超大单净额", "大单净额", "中单净额", "小单净额",
                           "主力流入资金", "主力流出资金",
                           "当日DDX", "当日DDY", "当日DDZ",
                           "3日DDX", "3日DDY",
                           "5日DDX", "5日DDY",
                           "10日DDX", "10日DDY",
                           "涨跌幅", "换手率", "振幅"):
                    if key not in data:  # 第一个 sheet 的优先
                        data[key] = _clean_value(row[1])

        return data

    except subprocess.TimeoutExpired:
        return {"error": "妙查脚本超时"}
    except Exception as e:
        return {"error": f"采集异常: {e}"}


def format_daily_log(data: dict) -> str:
    """格式化输出 markdown 格式的日志"""
    today = datetime.now().strftime("%Y-%m-%d")
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    if "error" in data:
        return f"""# {today} 每日跟踪

## 000629 攀钢钒钛

**采集失败**: {data['error']}

🕐 {now}
"""

    price = data.get("收盘价") or data.get("最新价")
    open_p = data.get("开盘")
    high = data.get("最高")
    low = data.get("最低")
    prev_close = data.get("昨收")
    main_net = data.get("主力净流入") or data.get("主力净额")
    super_net = data.get("超大单净额")
    big_net = data.get("大单净额")
    ddx = data.get("当日DDX")
    ddx3 = data.get("3日DDX")
    ddx5 = data.get("5日DDX")
    ddx10 = data.get("10日DDX")

    def _fmt(v, suffix="元"):
        if v is None:
            return "未知"
        if isinstance(v, (int, float)):
            if abs(v) >= 10000:
                return f"{v:,.0f}{suffix}"
            return f"{v}{suffix}"
        return str(v)

    price_s = _fmt(price)
    open_s = _fmt(open_p)
    high_s = _fmt(high)
    low_s = _fmt(low)
    prev_s = _fmt(prev_close)
    main_s = _fmt(main_net, suffix="万")
    super_s = _fmt(super_net, suffix="万")
    big_s = _fmt(big_net, suffix="万")
    ddx_s = _fmt(ddx, suffix="")
    ddx3_s = _fmt(ddx3, suffix="")
    ddx5_s = _fmt(ddx5, suffix="")
    ddx10_s = _fmt(ddx10, suffix="")

    # 涨幅估算(妙查今日数据未返回昨收,以 0% 显示)
    change_pct = "数据缺失"
    if isinstance(prev_close, (int, float)) and isinstance(price, (int, float)) and prev_close:
        change_pct = f"{(price - prev_close) / prev_close * 100:+.2f}%"

    log = f"""# {today} 每日跟踪

## 000629 攀钢钒钛

### 今日行情

| 项目 | 数据 |
|------|------|
| 收盘价 | **{price_s}**（{change_pct}） |
| 开盘 | {open_s} | 最高 | {high_s} | 最低 | {low_s} |
| 昨收 | {prev_s} |

**资金流向：**

| 资金类型 | 净额 |
|---------|------|
| 主力净流入 | **{main_s}** |
| 超大单净额 | {super_s} |
| 大单净额 | {big_s} |

**DDX/DDY 指标：**

| 指标 | 数值 |
|------|------|
| 当日DDX | {ddx_s} |
| 3日DDX | {ddx3_s} |
| 5日DDX | {ddx5_s} |
| 10日DDX | {ddx10_s} |

---

### 换庄假说跟踪

(由人工/AI 解读时补充)

🕐 {now}（数据来源：push2delay(资金流) + txstock(行情)，DDX 不提供）
"""
    return log


def write_memory(content: str, dry_run: bool = False) -> str:
    """写入 memory/YYYY-MM-DD.md,返回文件路径"""
    today = datetime.now().strftime("%Y-%m-%d")
    mem_dir = PROJECT_ROOT / "memory"
    mem_dir.mkdir(exist_ok=True)
    mem_file = mem_dir / f"{today}.md"

    if dry_run:
        return f"[DRY-RUN] {mem_file}"

    mem_file.write_text(content, encoding="utf-8")
    return str(mem_file)


def write_outbox(content: str, kind: str = "ftqq_daily", dry_run: bool = False) -> str:
    """写入 data/outbox/ 由小海转发飞书"""
    import uuid
    outbox_dir = PROJECT_ROOT / "astock-signal/data/outbox"
    outbox_dir.mkdir(parents=True, exist_ok=True)

    msg_id = uuid.uuid4().hex[:8]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{msg_id}.json"
    filepath = outbox_dir / filename

    if dry_run:
        return f"[DRY-RUN] {filepath}"

    payload = {
        "id": msg_id,
        "content": content,
        "msg_type": "markdown",
        "target": "ou_ee2947ff311d4978679c2a2d4433f62a",
        "kind": kind,
        "created_at": datetime.now().isoformat(),
        "status": "pending",
    }
    filepath.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                        encoding="utf-8")
    return str(filepath)


def main():
    parser = argparse.ArgumentParser(description="000629 每日自动采集")
    parser.add_argument("--dry-run", action="store_true", help="不写文件")
    parser.add_argument("--source", default="multi",
                        choices=["multi", "push2delay", "txstock", "miaochang"],
                        help="数据源: multi (默认, push2delay+txstock) / miaochang (旧版) / 单独源")
    args = parser.parse_args()

    print(f"[{datetime.now().isoformat()}] 开始采集 000629 (source={args.source}) ...")

    if args.source == "miaochang":
        data = fetch_miaochang_data()
    elif args.source == "push2delay":
        data = fetch_push2delay_data()
    elif args.source == "txstock":
        data = fetch_txstock_data()
    else:  # multi (default)
        data = fetch_000629_data()

    if "error" in data:
        print(f"❌ 采集失败: {data['error']}")
        sys.exit(1)

    print(f"✅ 数据拉取成功,字段数: {len(data)}")
    log = format_daily_log(data)

    mem_path = write_memory(log, dry_run=args.dry_run)
    print(f"{'🟡' if args.dry_run else '✅'} memory: {mem_path}")

    # outbox 内容是简化版(更适合飞书推送)
    # v6.11 fix: 用 "\n\n---" 切, 避免误中 markdown 表格分隔符 "|---|---|"
    summary = log.split("\n\n---")[0].strip()  # 取第一段 (行情 + 资金流)
    outbox_path = write_outbox(summary, kind="ftqq_daily",
                                dry_run=args.dry_run)
    print(f"{'🟡' if args.dry_run else '✅'} outbox: {outbox_path}")

    print(f"[{datetime.now().isoformat()}] 完成")


if __name__ == "__main__":
    main()
